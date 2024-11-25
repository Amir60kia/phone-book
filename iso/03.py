from fileinput import filename
# عالی کار میکند!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
import openpyxl
import os

# مسیر فایل اکسل خام
file_path = "01\EKFR-IT-003-00--1403.xlsx"
folder_path="01"
filename="EKFR-IT-003-00--1403.xlsx"
# شماره سطر و ستون برای سلول مورد نظر
row_number = 19
column_number = 9

# بارگذاری یا ایجاد فایل اکسل
if os.path.exists(file_path):
    workbook = openpyxl.load_workbook(file_path)
else:
    workbook = openpyxl.Workbook()  # ایجاد فایل جدید اگر فایل موجود نباشد

sheet = workbook.active
for i in range(32):
    print(i)
    # قرار دادن مقدار در سلول مورد نظر
    sheet.cell(row=row_number, column=column_number).value = "تاریخ : 1403.02."+str(i)+""

# ذخیره فایل
# workbook.save(file_path)
    new_filename = filename.replace("1403", "140302"+str(i))
    new_file_path = os.path.join(folder_path, new_filename)
    workbook.save(new_file_path)
    print("مقدار در سلول (19, 9) قرار داده شد و فایل ذخیره شد.")
